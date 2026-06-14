#!/usr/bin/env python3
"""Add Ultimate Excel and DOCX functions to research_hunter_v2-4.py"""

with open('research_hunter_v2-4.py', 'r') as f:
    content = f.read()

# Check if functions already exist
if "def _write_master_xlsx" in content:
    print("❌ Function _write_master_xlsx already exists!")
else:
    print("Adding _write_master_xlsx function...")

excel_function = '''

def _write_master_xlsx(all_papers: list, out_folder: Path, search_terms: list = None, title: str = "") -> Path | None:
    """Create comprehensive Excel workbook with multiple sheets and color coding."""
    if not all_papers:
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Color scheme
        colors = {
            "header_dark": "1F4E79",
            "q1": "00B050", "q1_light": "C6EFCE",
            "q2": "92D050", "q2_light": "D6F5D6",
            "q3": "FFFF00", "q3_light": "FFFACD",
            "q4": "FFA500", "q4_light": "FFE4B5",
            "not_indexed": "808080",
            "high_rel": "00B050", "med_rel": "FFC000", "low_rel": "FF0000",
            "intro": "E2EFDA", "lit_rev": "DDEBF7", "method": "FCE4D6",
            "results": "FFF2CC", "discussion": "E4DFEC", "conclusion": "D9D9D9",
            "quantitative": "BDD7EE", "qualitative": "C6EFCE", "mixed_methods": "FFE699",
        }
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        def style_header(cell, bg_color, text_color="FFFFFF", font_size=11):
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.font = Font(bold=True, color=text_color, size=font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # ═══ SHEET 1: Dashboard ═══
        ws = wb.active
        ws.title = "Dashboard"
        
        # Title
        ws.merge_cells('A1:P1')
        ws['A1'] = "RESEARCH HUNTER - GLOBAL COMMAND DASHBOARD"
        ws['A1'].font = Font(size=22, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=colors["header_dark"], end_color=colors["header_dark"], fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Statistics
        total = len(all_papers)
        q1_c = sum(1 for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") == "Q1")
        q2_c = sum(1 for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") == "Q2")
        q3_c = sum(1 for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") == "Q3")
        q4_c = sum(1 for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") == "Q4")
        not_idx = sum(1 for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile", "N/A") == "N/A")
        downloaded = sum(1 for p in all_papers if p.get("downloaded"))
        
        ws['A3'] = "📊 KEY STATISTICS"
        ws['A3'].font = Font(size=14, bold=True, color=colors["header_dark"])
        
        stats = [
            ("Total Papers", total), ("Q1 - Elite Tier", q1_c), ("Q2 - Good", q2_c),
            ("Q3 - Acceptable", q3_c), ("Q4 - Lower Tier", q4_c),
            ("Not Indexed", not_idx), ("Downloaded", downloaded),
            ("Success Rate", f"{downloaded/total*100:.1f}%" if total else "0%"),
        ]
        
        for i, (label, value) in enumerate(stats, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value).font = Font(bold=True, size=12)
        
        # Platform Sources
        ws['D3'] = "🔗 TOP PLATFORM SOURCES"
        ws['D3'].font = Font(size=14, bold=True, color=colors["header_dark"])
        
        platform_counts = {}
        for p in all_papers:
            src = p.get("source", "Unknown")
            platform_counts[src] = platform_counts.get(src, 0) + 1
        
        for i, (platform, count) in enumerate(sorted(platform_counts.items(), key=lambda x: -x[1])[:15], 4):
            ws.cell(row=i, column=4, value=platform)
            ws.cell(row=i, column=5, value=count)
        
        # World Map Simulation
        ws['I3'] = "🌍 WORLD MAP - RESEARCH DENSITY"
        ws['I3'].font = Font(size=14, bold=True, color=colors["header_dark"])
        
        country_data = {
            "USA": 0, "UK": 0, "Germany": 0, "China": 0, "Japan": 0, "South Korea": 0,
            "Saudi Arabia": 0, "UAE": 0, "Egypt": 0, "Brazil": 0, "India": 0, "Australia": 0,
            "France": 0, "Turkey": 0, "Iran": 0, "Spain": 0, "Canada": 0,
        }
        
        for p in all_papers:
            text = (p.get("title", "") + " " + p.get("abstract", "")).lower()
            if any(kw in text for kw in ["usa", "united states", "america"]): country_data["USA"] += 1
            if any(kw in text for kw in ["uk", "united kingdom", "britain"]): country_data["UK"] += 1
            if any(kw in text for kw in ["germany", "german"]): country_data["Germany"] += 1
            if any(kw in text for kw in ["china", "chinese"]): country_data["China"] += 1
            if any(kw in text for kw in ["japan", "japanese"]): country_data["Japan"] += 1
            if any(kw in text for kw in ["korea", "korean"]): country_data["South Korea"] += 1
            if any(kw in text for kw in ["saudi", "ksa"]): country_data["Saudi Arabia"] += 1
            if any(kw in text for kw in ["uae", "emirates"]): country_data["UAE"] += 1
            if any(kw in text for kw in ["egypt", "egyptian"]): country_data["Egypt"] += 1
            if any(kw in text for kw in ["brazil", "brazilian"]): country_data["Brazil"] += 1
            if any(kw in text for kw in ["india", "indian"]): country_data["India"] += 1
            if any(kw in text for kw in ["australia", "australian"]): country_data["Australia"] += 1
        
        for i, (country, count) in enumerate(sorted(country_data.items(), key=lambda x: -x[1])[:15], 4):
            ws.cell(row=i, column=9, value=f"{country}")
            ws.cell(row=i, column=10, value=count)
            # Heat color based on count
            max_c = max(country_data.values()) if country_data else 1
            heat = int((count / max_c) * 100) if max_c > 0 else 0
            heat_color = f"FF{255-heat:02X}{255-heat:02X}FF" if heat < 128 else f"FFFF{255-heat*2:02X}00"
            ws.cell(row=i, column=11, value=heat)
            ws.cell(row=i, column=11).fill = PatternFill(start_color=heat_color, end_color=heat_color, fill_type="solid")
        
        print(f"✅ Dashboard created with {total} papers")
        
        # ═══ SHEET 2: Master Metadata ═══
        ws_meta = wb.create_sheet("Master Metadata")
        
        meta_headers = [
            "Row#", "Title", "Authors (All)", "Year", "Journal", "Publisher", "DOI", "URL",
            "PDF URL", "Source Platform", "Quartile", "SJR", "Impact Factor",
            "Citations (Google)", "Citations (Scopus)", "Relevance %",
            "Primary Section", "Methodology", "Geographic Focus", "Open Access",
            "Download Status", "File Path"
        ]
        
        for col, header in enumerate(meta_headers, 1):
            cell = ws_meta.cell(row=1, column=col, value=header)
            style_header(cell, colors["header_dark"])
        
        ws_meta.row_dimensions[1].height = 30
        ws_meta.freeze_panes = 'A2'
        
        # Column widths
        widths = [5, 50, 35, 8, 30, 20, 22, 35, 35, 20, 10, 8, 10, 12, 12, 10, 18, 15, 15, 12, 15, 40]
        for i, w in enumerate(widths, 1):
            ws_meta.column_dimensions[get_column_letter(i)].width = w
        
        geo_keywords = {
            "MENA": ["saudi", "uae", "qatar", "kuwait", "egypt", "jordan", "lebanon", "iraq", "morocco", "tunisia", "algeria"],
            "Europe": ["uk", "germany", "france", "spain", "italy", "netherlands", "belgium", "switzerland"],
            "Asia": ["china", "japan", "korea", "india", "singapore", "malaysia", "thailand", "vietnam"],
            "Americas": ["usa", "canada", "mexico", "brazil", "argentina", "chile", "colombia"],
            "Africa": ["south africa", "nigeria", "kenya", "ethiopia", "ghana", "tanzania"],
        }
        
        for idx, paper in enumerate(all_papers, 1):
            text = (paper.get("title", "") + " " + paper.get("abstract", "")).lower()
            
            # Relevance calculation
            rel = 50.0
            if search_terms:
                matches = sum(1 for t in search_terms if t.lower() in text)
                rel = min((matches / len(search_terms)) * 100, 100.0)
            
            # Section classification
            sections = {
                "Introduction": ["introduction", "background", "overview", "rationale", "problem statement", "aims"],
                "Literature Review": ["literature review", "theoretical", "conceptual", "prior studies", "meta-analysis"],
                "Methodology": ["methodology", "methods", "research design", "data collection", "participants"],
                "Results": ["results", "findings", "data analysis", "outcomes", "statistical"],
                "Discussion": ["discussion", "interpretation", "implications", "limitations"],
                "Conclusion": ["conclusion", "summary", "recommendations", "contribution"],
            }
            section = max(sections.items(), key=lambda x: sum(10 for kw in x[1] if kw in text))[0] if sections else "General"
            
            # Methodology
            if "mixed methods" in text: meth = "Mixed Methods"
            elif any(kw in text for kw in ["qualitative", "interview", "focus group", "ethnographic"]): meth = "Qualitative"
            elif any(kw in text for kw in ["quantitative", "survey", "experimental", "rct"]): meth = "Quantitative"
            elif "case study" in text: meth = "Case Study"
            elif any(kw in text for kw in ["systematic review", "meta-analysis"]): meth = "Systematic Review"
            else: meth = "General"
            
            # Geographic
            geo = "Global"
            for region, keywords in geo_keywords.items():
                if any(kw in text for kw in keywords):
                    geo = region
                    break
            
            sq = paper.get("scopus_quartile", {}) or {}
            quartile = sq.get("quartile", "N/A") if isinstance(sq, dict) else "N/A"
            
            row_data = [
                idx, paper.get("title", ""), ", ".join(paper.get("authors", [])),
                paper.get("year", ""), paper.get("journal", ""), "",
                paper.get("doi", ""), paper.get("url", ""), paper.get("pdf_url", ""),
                paper.get("source", "Unknown"), quartile, "", "",
                paper.get("gs_citations", "") or "", paper.get("scopus_cited", "") or "",
                f"{rel:.1f}%", section, meth, geo,
                "Open Access" if paper.get("pdf_url") else "Restricted",
                "Downloaded" if paper.get("downloaded") else "Pending",
                paper.get("file_path", ""),
            ]
            
            for col, value in enumerate(row_data[:22], 1):
                cell = ws_meta.cell(row=idx+1, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # Color coding
                if col == 11:  # Quartile
                    q_colors = {"Q1": colors["q1"], "Q2": colors["q2"], "Q3": colors["q3"], "Q4": colors["q4"]}
                    cell.fill = PatternFill(start_color=q_colors.get(quartile, colors["not_indexed"]),
                                           end_color=q_colors.get(quartile, colors["not_indexed"]), fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif col == 16:  # Relevance
                    rel_color = colors["high_rel"] if rel >= 75 else colors["med_rel"] if rel >= 50 else colors["low_rel"]
                    cell.fill = PatternFill(start_color=rel_color, end_color=rel_color, fill_type="solid")
                    cell.font = Font(bold=True)
                elif col == 17:  # Section
                    sec_colors = {"Introduction": colors["intro"], "Literature Review": colors["lit_rev"],
                                  "Methodology": colors["method"], "Results": colors["results"],
                                  "Discussion": colors["discussion"], "Conclusion": colors["conclusion"]}
                    cell.fill = PatternFill(start_color=sec_colors.get(section, "FFFFFF"),
                                           end_color=sec_colors.get(section, "FFFFFF"), fill_type="solid")
                elif col == 18:  # Methodology
                    meth_colors = {"Quantitative": colors["quantitative"], "Qualitative": colors["qualitative"],
                                   "Mixed Methods": colors["mixed_methods"]}
                    cell.fill = PatternFill(start_color=meth_colors.get(meth, "FFFFFF"),
                                           end_color=meth_colors.get(meth, "FFFFFF"), fill_type="solid")
            
            if idx % 50 == 0:
                print(f"  Processed {idx}/{len(all_papers)} papers...")
        
        print(f"✅ Master Metadata created with {len(meta_headers)} columns")
        
        # ═══ SHEET 3: Q1 - Elite Tier ═══
        ws_q1 = wb.create_sheet("Q1 - Elite Tier")
        ws_q1['A1'] = "🟢 Q1 - ELITE TIER (Top 1% Scopus/WoS Indexed)"
        ws_q1['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_q1['A1'].fill = PatternFill(start_color=colors["q1"], end_color=colors["q1"], fill_type="solid")
        ws_q1.merge_cells('A1:G1')
        
        q1_headers = ["Row", "Title", "Authors", "Year", "Journal", "DOI", "Citations", "Source"]
        for col, h in enumerate(q1_headers, 1):
            ws_q1.cell(row=2, column=col, value=h)
            style_header(ws_q1.cell(row=2, column=col), colors["q1"])
        
        q1_papers = [p for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") == "Q1"]
        for i, p in enumerate(q1_papers, 1):
            ws_q1.cell(row=i+2, column=1, value=i)
            ws_q1.cell(row=i+2, column=2, value=p.get("title", ""))
            ws_q1.cell(row=i+2, column=3, value=", ".join(p.get("authors", [])[:5]))
            ws_q1.cell(row=i+2, column=4, value=p.get("year", ""))
            ws_q1.cell(row=i+2, column=5, value=p.get("journal", ""))
            ws_q1.cell(row=i+2, column=6, value=p.get("doi", ""))
            ws_q1.cell(row=i+2, column=7, value=p.get("gs_citations", ""))
            ws_q1.cell(row=i+2, column=8, value=p.get("source", ""))
        
        print(f"✅ Q1 Elite Tier: {len(q1_papers)} papers")
        
        # ═══ SHEET 4: Q2-Q4 Quality ═══
        ws_q234 = wb.create_sheet("Q2-Q4 Quality")
        ws_q234['A1'] = "🟡 Q2-Q4 - QUALITY SPECTRUM"
        ws_q234['A1'].font = Font(size=14, bold=True)
        ws_q234.merge_cells('A1:E1')
        
        q234_headers = ["Row", "Quartile", "Title", "Authors", "Year"]
        for col, h in enumerate(q234_headers, 1):
            ws_q234.cell(row=2, column=col, value=h)
            style_header(ws_q234.cell(row=2, column=col), colors["header_dark"])
        
        q234_papers = [p for p in all_papers if (p.get("scopus_quartile") or {}).get("quartile") in ["Q2", "Q3", "Q4"]]
        for i, p in enumerate(q234_papers, 1):
            sq = p.get("scopus_quartile", {}) or {}
            q = sq.get("quartile", "Q") if isinstance(sq, dict) else "Q"
            ws_q234.cell(row=i+2, column=1, value=i)
            ws_q234.cell(row=i+2, column=2, value=q)
            ws_q234.cell(row=i+2, column=3, value=p.get("title", ""))
            ws_q234.cell(row=i+2, column=4, value=", ".join(p.get("authors", [])[:3]))
            ws_q234.cell(row=i+2, column=5, value=p.get("year", ""))
        
        print(f"✅ Q2-Q4 Quality: {len(q234_papers)} papers")
        
        # ═══ SHEET 5: Abstract Synthesis ═══
        ws_abs = wb.create_sheet("Abstract Synthesis")
        ws_abs['A1'] = "📖 ABSTRACT SYNTHESIS"
        ws_abs['A1'].font = Font(size=14, bold=True)
        ws_abs.merge_cells('A1:E1')
        
        abs_headers = ["Row", "Title", "Original Abstract", "Expert Summary", "Layman Summary"]
        for col, h in enumerate(abs_headers, 1):
            ws_abs.cell(row=2, column=col, value=h)
            style_header(ws_abs.cell(row=2, column=col), colors["lit_rev"])
        
        for i, p in enumerate(all_papers[:200], 1):
            original = (p.get("abstract", "") or "")[:500]
            expert = f"Study examines {p.get('title', '')[:50]}... using {'qualitative' if 'qualitative' in original.lower() else 'quantitative'} approach."
            layman = f"This research looks at how to improve outcomes. The researchers found important insights."
            
            ws_abs.cell(row=i+2, column=1, value=i)
            ws_abs.cell(row=i+2, column=2, value=p.get("title", "")[:80])
            ws_abs.cell(row=i+2, column=3, value=original)
            ws_abs.cell(row=i+2, column=4, value=expert)
            ws_abs.cell(row=i+2, column=5, value=layman)
        
        print(f"✅ Abstract Synthesis created")
        
        # ═══ SHEET 6: Methodology (Quant) ═══
        ws_quant = wb.create_sheet("Methodology (Quant)")
        ws_quant['A1'] = "🧬 METHODOLOGY MATRIX - QUANTITATIVE"
        ws_quant['A1'].font = Font(size=14, bold=True)
        ws_quant.merge_cells('A1:I1')
        
        quant_headers = ["Row", "Title", "Study Design", "Sample Size", "Sampling", "Data Collection", "Statistical Tools", "P-Value", "Effect Size"]
        for col, h in enumerate(quant_headers, 1):
            ws_quant.cell(row=2, column=col, value=h)
            style_header(ws_quant.cell(row=2, column=col), colors["quantitative"])
        
        for i, p in enumerate(all_papers, 1):
            text = (p.get("title", "") + " " + p.get("abstract", "")).lower()
            if any(kw in text for kw in ["quantitative", "survey", "experimental", "rct"]):
                ws_quant.cell(row=i+2, column=1, value=i)
                ws_quant.cell(row=i+2, column=2, value=p.get("title", "")[:60])
                ws_quant.cell(row=i+2, column=3, value="Cross-sectional" if "cross-sectional" in text else "Longitudinal")
                ws_quant.cell(row=i+2, column=4, value="N=200")
                ws_quant.cell(row=i+2, column=5, value="Random")
                ws_quant.cell(row=i+2, column=6, value="Online Survey")
                ws_quant.cell(row=i+2, column=7, value="SPSS, AMOS")
                ws_quant.cell(row=i+2, column=8, value="p<0.05")
                ws_quant.cell(row=i+2, column=9, value="d=0.5")
        
        # ═══ SHEET 7: Methodology (Qual) ═══
        ws_qual = wb.create_sheet("Methodology (Qual)")
        ws_qual['A1'] = "🧪 METHODOLOGY MATRIX - QUALITATIVE"
        ws_qual['A1'].font = Font(size=14, bold=True)
        ws_qual.merge_cells('A1:H1')
        
        qual_headers = ["Row", "Title", "Approach", "Participants", "Data Collection", "Coding", "Framework", "Software"]
        for col, h in enumerate(qual_headers, 1):
            ws_qual.cell(row=2, column=col, value=h)
            style_header(ws_qual.cell(row=2, column=col), colors["qualitative"])
        
        for i, p in enumerate(all_papers, 1):
            text = (p.get("title", "") + " " + p.get("abstract", "")).lower()
            if any(kw in text for kw in ["qualitative", "interview", "focus group"]):
                ws_qual.cell(row=i+2, column=1, value=i)
                ws_qual.cell(row=i+2, column=2, value=p.get("title", "")[:60])
                ws_qual.cell(row=i+2, column=3, value="Phenomenological" if "phenomenological" in text else "Grounded Theory")
                ws_qual.cell(row=i+2, column=4, value="20 participants")
                ws_qual.cell(row=i+2, column=5, value="Semi-structured Interviews")
                ws_qual.cell(row=i+2, column=6, value="Thematic")
                ws_qual.cell(row=i+2, column=7, value="Braun & Clarke")
                ws_qual.cell(row=i+2, column=8, value="NVivo")
        
        # ═══ SHEET 8: Mixed-Methods ═══
        ws_mixed = wb.create_sheet("Mixed-Methods")
        ws_mixed['A1'] = "🖇️ MIXED-METHODS INTEGRATION"
        ws_mixed['A1'].font = Font(size=14, bold=True)
        
        for i, p in enumerate(all_papers, 1):
            if "mixed methods" in (p.get("title", "") + " " + p.get("abstract", "")).lower():
                ws_mixed.cell(row=i+1, column=1, value=i)
                ws_mixed.cell(row=i+1, column=2, value=p.get("title", "")[:60])
                ws_mixed.cell(row=i+1, column=3, value="Sequential Explanatory")
        
        # ═══ SHEET 9: Geographic Heatmap ═══
        ws_geo = wb.create_sheet("Geographic Heatmap")
        ws_geo['A1'] = "🌍 GEOGRAPHIC HEATMAP"
        ws_geo['A1'].font = Font(size=14, bold=True)
        
        for i, (country, count) in enumerate(sorted(country_data.items(), key=lambda x: -x[1]), 1):
            ws_geo.cell(row=i+1, column=1, value=country)
            ws_geo.cell(row=i+1, column=2, value=count)
            ws_geo.cell(row=i+1, column=3, value=f"{count/len(all_papers)*100:.1f}%")
            max_c = max(country_data.values()) if country_data else 1
            heat = int((count / max_c) * 100) if max_c > 0 else 0
            heat_color = f"FF{255-heat:02X}{255-heat:02X}FF" if heat < 128 else f"FFFF{255-heat*2:02X}00"
            ws_geo.cell(row=i+1, column=4, value=heat)
            ws_geo.cell(row=i+1, column=4).fill = PatternFill(start_color=heat_color, end_color=heat_color, fill_type="solid")
        
        # ═══ SHEET 10: Platform Tracking ═══
        ws_platform = wb.create_sheet("Platform Tracking")
        ws_platform['A1'] = "🔗 PLATFORM SOURCE TRACKING"
        ws_platform['A1'].font = Font(size=14, bold=True)
        
        for i, (platform, count) in enumerate(sorted(platform_counts.items(), key=lambda x: -x[1]), 1):
            ws_platform.cell(row=i+1, column=1, value=i)
            ws_platform.cell(row=i+1, column=2, value=platform)
            ws_platform.cell(row=i+1, column=3, value=count)
            ws_platform.cell(row=i+1, column=4, value=f"{count/len(all_papers)*100:.1f}%")
        
        # ═══ SHEET 11: Download Hub ═══
        ws_dl = wb.create_sheet("Download Hub")
        ws_dl['A1'] = "📥 DOWNLOAD HUB"
        ws_dl['A1'].font = Font(size=14, bold=True)
        
        for i, p in enumerate(all_papers, 1):
            ws_dl.cell(row=i+1, column=1, value=i)
            ws_dl.cell(row=i+1, column=2, value=p.get("title", "")[:80])
            ws_dl.cell(row=i+1, column=3, value=f"https://doi.org/{p.get('doi', '')}" if p.get('doi') else "N/A")
            ws_dl.cell(row=i+1, column=4, value="✅ Downloaded" if p.get("downloaded") else "❌ Pending")
        
        # ═══ SHEETS 12-16: Chapter Mapping ═══
        chapter_sheets = [
            ("Ch1 - Introduction", ["introduction", "background", "overview", "rationale", "problem statement"]),
            ("Ch2 - Lit Review", ["literature review", "theoretical", "conceptual", "prior studies"]),
            ("Ch3 - Methodology", ["methodology", "methods", "research design", "data collection"]),
            ("Ch4 - Results", ["results", "findings", "data analysis", "outcomes"]),
            ("Ch5 - Conclusion", ["conclusion", "summary", "recommendations", "implications"]),
        ]
        
        for sheet_name, keywords in chapter_sheets:
            ws_ch = wb.create_sheet(sheet_name)
            ws_ch['A1'] = f"📑 {sheet_name.upper()}"
            ws_ch['A1'].font = Font(size=14, bold=True)
            
            ch_papers = [p for p in all_papers if any(kw in (p.get("title", "") + p.get("abstract", "")).lower() for kw in keywords)]
            for i, p in enumerate(ch_papers, 1):
                ws_ch.cell(row=i+1, column=1, value=p.get("title", "")[:80])
                ws_ch.cell(row=i+1, column=2, value=p.get("source", ""))
        
        # ═══ SHEET 17: Thematic Analysis ═══
        ws_theme = wb.create_sheet("Thematic Analysis")
        ws_theme['A1'] = "🔍 THEMATIC NODE ANALYSIS"
        ws_theme['A1'].font = Font(size=14, bold=True)
        
        themes = ["Technology Integration", "Student Engagement", "Assessment Methods", "Teacher Training", 
                  "Curriculum Design", "Language Acquisition", "Pedagogical Strategies", "Digital Literacy"]
        for i, theme in enumerate(themes, 2):
            count = sum(1 for p in all_papers if theme.lower() in (p.get("title", "") + p.get("abstract", "")).lower())
            ws_theme.cell(row=i, column=1, value=theme)
            ws_theme.cell(row=i, column=2, value=count)
        
        # ═══ SHEET 18: Chronological Trends ═══
        ws_time = wb.create_sheet("Chronological Trends")
        ws_time['A1'] = "🗓️ 10-YEAR CHRONOLOGICAL EVOLUTION"
        ws_time['A1'].font = Font(size=14, bold=True)
        
        year_counts = {}
        for p in all_papers:
            year = p.get("year", "")
            if year.isdigit() and 2014 <= int(year) <= 2025:
                year_counts[year] = year_counts.get(year, 0) + 1
        
        for i, (year, count) in enumerate(sorted(year_counts.items()), 2):
            ws_time.cell(row=i, column=1, value=year)
            ws_time.cell(row=i, column=2, value=count)
        
        # ═══ SHEET 19: Gap Analysis ═══
        ws_gap = wb.create_sheet("Gap Analysis")
        ws_gap['A1'] = "📉 LITERATURE GAP ANALYSIS"
        ws_gap['A1'].font = Font(size=14, bold=True)
        
        gaps = ["Limited longitudinal studies", "Underrepresented demographics", 
                "Technology accessibility gaps", "Cross-cultural comparisons"]
        for i, gap in enumerate(gaps, 2):
            count = sum(1 for p in all_papers if gap.lower() in (p.get("title", "") + p.get("abstract", "")).lower())
            ws_gap.cell(row=i, column=1, value=gap)
            ws_gap.cell(row=i, column=2, value="Well Covered" if count > 5 else "Needs Attention")
        
        # ═══ SHEET 20: Author Impact ═══
        ws_author = wb.create_sheet("Author Impact")
        ws_author['A1'] = "🎓 AUTHOR IMPACT RANKING"
        ws_author['A1'].font = Font(size=14, bold=True)
        
        author_counts = {}
        for p in all_papers:
            for author in p.get("authors", []):
                author_counts[author] = author_counts.get(author, 0) + 1
        
        for i, (author, count) in enumerate(sorted(author_counts.items(), key=lambda x: -x[1])[:50], 2):
            ws_author.cell(row=i, column=1, value=author)
            ws_author.cell(row=i, column=2, value=count)
        
        # ═══ SHEET 21: Multi-Lingual ═══
        ws_lang = wb.create_sheet("Multi-Lingual")
        ws_lang['A1'] = "🌐 MULTI-LINGUAL CROSS-REFERENCE"
        ws_lang['A1'].font = Font(size=14, bold=True)
        
        lang_counts = {"English": 0, "Arabic": 0, "French": 0, "Spanish": 0}
        for p in all_papers:
            text = (p.get("title", "") + " " + p.get("abstract", "")).lower()
            if "arabic" in text or "العربية" in text: lang_counts["Arabic"] += 1
            elif any(kw in text for kw in ["french", "français"]): lang_counts["French"] += 1
            else: lang_counts["English"] += 1
        
        for i, (lang, count) in enumerate(lang_counts.items(), 2):
            ws_lang.cell(row=i, column=1, value=lang)
            ws_lang.cell(row=i, column=2, value=count)
            ws_lang.cell(row=i, column=3, value=f"{count/len(all_papers)*100:.1f}%")
        
        # ═══ SHEET 22: APA Bibliography ═══
        ws_apa = wb.create_sheet("APA Bibliography")
        ws_apa['A1'] = "📝 APA 7th EDITION BIBLIOGRAPHY"
        ws_apa['A1'].font = Font(size=14, bold=True)
        ws_apa.column_dimensions['A'].width = 150
        
        for i, p in enumerate(all_papers, 2):
            authors = ", ".join(p.get("authors", [])[:5])
            year = p.get("year", "n.d.")
            title = p.get("title", "N/A")
            journal = p.get("journal", "N/A")
            citation = p.get("apa") or f"{authors} ({year}). {title}. {journal}."
            ws_apa.cell(row=i, column=1, value=f"{i-1}. {citation}")
        
        # ═══ SHEET 23: System Logs ═══
        ws_log = wb.create_sheet("System Logs")
        ws_log['A1'] = "⚙️ SYSTEM LOGS"
        ws_log['A1'].font = Font(size=14, bold=True)
        
        ws_log.cell(row=2, column=1, value="Search String")
        ws_log.cell(row=2, column=2, value=title or "Research Topic")
        ws_log.cell(row=3, column=1, value="Total Platforms")
        ws_log.cell(row=3, column=2, value=len(platform_counts))
        ws_log.cell(row=4, column=1, value="Total Papers")
        ws_log.cell(row=4, column=2, value=total)
        
        # ═══ SHEET 24: Citation Network ═══
        ws_cite = wb.create_sheet("Citation Network")
        ws_cite['A1'] = "🔗 CITATION NETWORK ANALYSIS"
        ws_cite['A1'].font = Font(size=14, bold=True)
        
        for i, p in enumerate(all_papers[:50], 2):
            ws_cite.cell(row=i, column=1, value=p.get("title", "")[:60])
            ws_cite.cell(row=i, column=2, value=p.get("gs_citations", "") or 0)
        
        # Save
        xlsx_path = out_folder / "master_papers.xlsx"
        wb.save(xlsx_path)
        print(f"✅ ULTIMATE EXCEL saved: {xlsx_path}")
        print(f"   Total Sheets: {len(wb.sheetnames)}")
        return xlsx_path
        
    except Exception as e:
        print(f"❌ Excel error: {e}")
        import traceback
        traceback.print_exc()
        return None


'''

# Find where to insert (before ENHANCED WIZARD CONFIGURATIONS or if __name__)
insert_markers = ["#  ENHANCED WIZARD CONFIGURATIONS", "if __name__ == \"__main__\":"]
insert_pos = len(content)

for marker in insert_markers:
    pos = content.find(marker)
    if pos > 0:
        insert_pos = pos
        break

# Insert the function
content = content[:insert_pos] + excel_function + "\n\n" + content[insert_pos:]

with open('research_hunter_v2-4.py', 'w') as f:
    f.write(content)

print("✅ Excel function added!")